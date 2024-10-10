from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuração do navegador
chrome_options = Options()
chrome_options.add_argument("--headless")  # Executa o Chrome em modo headless (sem interface gráfica)
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# Usa o WebDriver Manager para instalar a versão correta do ChromeDriver automaticamente
service = Service(ChromeDriverManager().install())

# Inicializa o WebDriver do Chrome
driver = webdriver.Chrome(service=service, options=chrome_options)

# URL de busca no YouTube
query = 'Músicas contra ansiedade'
url = f'https://www.youtube.com/results?search_query={query}'

# Acessa o YouTube
driver.get(url)

# Usar espera explícita para garantir que os elementos estão carregados
wait = WebDriverWait(driver, 15)

# Procurar os containers que contêm os vídeos
containers = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//ytd-video-renderer')))

# Listas para armazenar os dados
video_titles = []
video_urls = []

# Extrair e mostrar os vídeos encontrados
for container in containers:
    try:
        video_title = container.find_element(By.XPATH, './/*[@id="video-title"]').get_attribute('title')
        video_url = container.find_element(By.XPATH, './/*[@id="video-title"]').get_attribute('href')

        if video_title and video_url:
            video_titles.append(video_title)
            video_urls.append(video_url)
    except Exception as e:
        continue

# Fecha o WebDriver
driver.quit()

# Criar um DataFrame com os dados coletados
df = pd.DataFrame({
    'Título do Vídeo': video_titles,
    'URL': video_urls
})

# Exibe o DataFrame no console para visualização
print(df)

# Criar um workbook e uma sheet
wb = Workbook()
ws = wb.active
ws.title = "Vídeos de Relaxamento"

# Adicionar cabeçalhos
ws.append(['Título do Vídeo', 'URL'])

# Adicionar dados
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# Ajustar a largura das colunas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Obter a letra da coluna
    for cell in col:
        try:  # Encontrar o valor mais longo na coluna
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 5)  # Aumenta o fator de ajuste para comportar URLs longas
    ws.column_dimensions[column].width = adjusted_width

# Definir bordas finas para todas as células
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Estilo para o cabeçalho
for cell in ws["1:1"]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="0000CC", end_color="0000CC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border  # Aplicar borda ao cabeçalho

# Aplicar alinhamento, bordas e cor alternada às linhas (ignorando o cabeçalho)
for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
    fill_color = PatternFill(start_color="C9F5FF", end_color="C9F5FF", fill_type="solid") if row_index % 2 != 0 else PatternFill(start_color="eaeaea", end_color="eaeaea", fill_type="solid")

    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = fill_color
        cell.border = thin_border  # Aplicar borda em cada célula

# Salvar o arquivo Excel
output_path = r'C:\Users\colet\OneDrive\Área de Trabalho\PDSI\WebScraping\videos_relaxamento_youtube.xlsx'
wb.save(output_path)

print(f'Dados salvos com sucesso no arquivo {output_path}')
